"""导入 tk.xlsx -> src/data/questions.json"""
import json
from pathlib import Path

from openpyxl import load_workbook

SRC = Path("tk.xlsx")
OUT = Path("src/data")
OUT.mkdir(parents=True, exist_ok=True)

wb = load_workbook(SRC, data_only=True)
TYPE_MAP = {"单选题": "single", "多选题": "multiple", "填空题": "fill", "判断题": "bool"}


def clean_choice_answer(raw_answer, raw_analysis, options, qtype):
    raw = str(raw_answer).strip().upper() if raw_answer else ""
    analysis = str(raw_analysis).strip() if raw_analysis else ""

    if "解析" in raw:
        answer_part, analysis_part = raw.split("解析", 1)
        raw = answer_part.strip(" ：:，,;；")
        analysis_from_answer = analysis_part.strip(" ：:，,;；")
        if analysis_from_answer and not analysis:
            analysis = analysis_from_answer

    valid = set(options.keys())
    if qtype == "single":
        for ch in raw:
            if ch in valid:
                return ch, analysis
    else:
        cleaned = "".join(ch for ch in raw if ch in valid)
        if cleaned:
            return "".join(dict.fromkeys(cleaned)), analysis

    # Fallback for rows where the answer cell contains only analysis text.
    combined = " ".join([str(raw_answer or ""), analysis])
    for key, value in options.items():
        text = str(value).strip()
        if text and f"{text}并不是" in combined:
            return key, analysis

    return raw, analysis


all_questions = []
for sheet_name in ["单选题", "多选题", "填空题", "判断题"]:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for i, row in enumerate(rows):
        qtype = TYPE_MAP[sheet_name]
        order = i + 1

        if qtype in ("single", "multiple"):
            seq, title, a, b, c, d, e, answer, analysis = row
            options = {
                k: str(v).strip()
                for k, v in {"A": a, "B": b, "C": c, "D": d, "E": e}.items()
                if v is not None and str(v).strip()
            }
            ans, analysis_text = clean_choice_answer(answer, analysis, options, qtype)
        elif qtype == "fill":
            seq, title, answer, analysis = row
            options = {}
            ans = str(answer).strip() if answer else ""
            analysis_text = str(analysis).strip() if analysis else ""
        elif qtype == "bool":
            seq, title, answer, analysis = row
            options = {"A": "对", "B": "错"}
            raw = str(answer).strip()
            ans = "A" if raw in ("对", "√", "T", "t", "True", "true", "TRUE") else "B"
            analysis_text = str(analysis).strip() if analysis else ""

        q = {
            "id": f"{qtype}-{order}",
            "type": qtype,
            "order": order,
            "title": str(title).strip() if title else "",
            "options": options,
            "answer": ans,
            "analysis": analysis_text,
        }
        all_questions.append(q)

with open(OUT / "questions.json", "w", encoding="utf-8") as f:
    json.dump(all_questions, f, ensure_ascii=False, indent=2)

print(f"Imported {len(all_questions)} questions -> {OUT / 'questions.json'}")
